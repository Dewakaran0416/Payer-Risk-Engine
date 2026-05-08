#!/usr/bin/env python3
"""
=============================================================================
  PAYER CLAIMS RISK ENGINE  v4.0
  Features:
    1. AR Days Saved — per risk score and action category
    2. Estimated Success Rate + Collection Projection ($)
    3. No-denial handling — uses prior action notes from input
    4. Auto-generated HTML Dashboard from output data

  Usage (CLI):
    python payer_risk_engine.py \
        --train  data/training_data.csv \
        --input  data/input_claims.csv \
        --output results/output.csv \
        --dashboard results/dashboard.html \
        --avg-claim 850

  Usage (API):
    from payer_risk_engine import PayerRiskEngine
    engine = PayerRiskEngine()
    engine.train("training_data.csv")
    results = engine.predict("input_claims.csv")
    engine.generate_dashboard(results, "dashboard.html")
    engine.export(results, "output.csv")
=============================================================================
"""

import argparse, json, logging, sys, time
from pathlib import Path
from typing import Optional
import pandas as pd
import numpy as np

logging.basicConfig(format="%(asctime)s  %(levelname)-8s  %(message)s",
                    datefmt="%H:%M:%S", level=logging.INFO)
log = logging.getLogger("PayerRiskEngine")

# ── Constants ──────────────────────────────────────────────────────────────

AR_MAP = {
    "modifier":        {"baseline": 90,  "after": 25,  "saved": 65},
    "medical_records": {"baseline": 180, "after": 70,  "saved": 110},
    "adjust":          {"baseline": 200, "after": 10,  "saved": 190},
    "auth":            {"baseline": 120, "after": 40,  "saved": 80},
    "appeal":          {"baseline": 210, "after": 85,  "saved": 125},
    "bill_patient":    {"baseline": 60,  "after": 18,  "saved": 42},
    "timely_filing":   {"baseline": 30,  "after": 5,   "saved": 25},
    "resubmit":        {"baseline": 75,  "after": 25,  "saved": 50},
    "default":         {"baseline": 120, "after": 55,  "saved": 65},
}

SUCCESS_RATE = {
    "CO-50": 0.88, "CO-197": 0.64, "CO-18": 0.25, "CO-30": 0.20,
    "CO-4":  0.82, "CO-11":  0.75, "CO-22": 0.70, "CO-29": 0.05,
    "CO-45": 0.15, "CO-97":  0.60, "CO-119":0.20, "CO-234":0.55,
    "CO-242":0.58, "PR-1":   0.95, "PR-2":  0.98, "PR-3":  0.97,
    "PR-204":0.30, "DEFAULT":0.55,
}

CONF_MULT = {"High": 1.00, "Medium": 0.88, "Low": 0.72}

SCORE_MULT = {(13,20):1.15, (9,13):1.00, (5,9):0.85, (0,5):0.70}

DENIAL_DEFAULTS = {
    "CO-4":  {"action":"Obtain prior authorization and resubmit","days":90},
    "CO-11": {"action":"Correct Diagnosis code and resubmit or appeal","days":80},
    "CO-18": {"action":"Adjust the Code — non-billable per coding team","days":180},
    "CO-22": {"action":"Coordinate benefits with secondary payer","days":60},
    "CO-29": {"action":"Write off — timely filing limit exceeded","days":30},
    "CO-30": {"action":"Adjust the Code — non-billable per payer policy","days":200},
    "CO-45": {"action":"Contractual adjustment — write off difference","days":45},
    "CO-50": {"action":"Add modifier (25/51/59) and resubmit the claim","days":75},
    "CO-97": {"action":"Review bundling edits; adjust or resubmit","days":60},
    "CO-119":{"action":"Adjust — benefit maximum reached; bill patient","days":45},
    "CO-197":{"action":"Submit with Medical Records and appeal if denied","days":160},
    "CO-234":{"action":"Appeal with clinical notes and MR documentation","days":180},
    "CO-242":{"action":"Appeal with medical necessity documentation","days":90},
    "PR-1":  {"action":"Bill patient — deductible not yet met","days":14},
    "PR-2":  {"action":"Bill patient — copay due","days":7},
    "PR-3":  {"action":"Bill patient — coinsurance due","days":7},
    "PR-204":{"action":"Bill patient or adjust — no benefit","days":30},
}

PAYER_FAMILIES = {
    "bcbs":     ["bcbs","blue cross","blue shield"],
    "aetna":    ["aetna"],
    "uhc":      ["uhc","united health","unitedhealthcare"],
    "medicare": ["medicare"],
    "medicaid": ["medicaid"],
    "cigna":    ["cigna"],
    "humana":   ["humana"],
    "molina":   ["molina"],
    "anthem":   ["anthem"],
    "centene":  ["centene"],
}

TRAIN_ALIASES = {
    "payer":  ["payer_name","payer","payername"],
    "cpt":    ["cpt_code","cpt","cptcode","procedure_code"],
    "dx":     ["dx_code","dx","dxcode","diagnosis_code"],
    "days":   ["resolution_days","days","resolutiondays"],
    "denial": ["denial_reason","denial","denial_code","denialcode"],
    "action": ["final_action","action","finalaction","recommended_action"],
    "notes":  ["historical_notes","notes","historicalnotes","activity"],
}

INPUT_ALIASES = {
    "payer":    ["payer_name","payer","payername","insurance"],
    "cpt":      ["cpt_code","cpt","cptcode"],
    "dx":       ["dx_code","dx","dxcode","diagnosis_code"],
    "denial":   ["denial_reason","denial","denial_code","denialcode"],
    "notes":    ["notes","activity","prior_action","previous_action","history","comment"],
    "amount":   ["charge_amount","amount","billed_amount","charge","claim_amount","fee"],
    "claim_id": ["claim_id","claimid","claim_number","ref"],
}

# ── Helpers ────────────────────────────────────────────────────────────────

def ns(v) -> str:
    return "" if (v is None or (isinstance(v, float) and np.isnan(v))) else str(v).strip()

def payer_fam(p: str) -> str:
    pl = p.lower()
    for fam, kws in PAYER_FAMILIES.items():
        if any(k in pl for k in kws): return fam
    return pl[:5]

def find_col(df, aliases, required=True):
    cols = {c.lower().replace(" ","_").replace("-","_"): c for c in df.columns}
    for a in aliases:
        k = a.lower().replace(" ","_")
        if k in cols: return cols[k]
    for a in aliases:
        for ck, co in cols.items():
            if a.replace("_","") in ck.replace("_",""): return co
    if required:
        raise ValueError(f"Cannot find column. Expected one of {aliases}. Got: {list(df.columns)}")
    return None

def load_csv_or_excel(path, aliases_map):
    p = Path(path)
    df = pd.read_excel(path, dtype=str) if p.suffix.lower() in [".xlsx",".xls"] \
         else pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    df.columns = df.columns.str.strip()
    mapped = {}
    for std, aliases in aliases_map.items():
        req = std not in ["notes","amount","claim_id"]
        col = find_col(df, aliases, required=req)
        mapped[std] = df[col].fillna("").str.strip() if col else pd.Series([""] * len(df))
    result = pd.DataFrame(mapped)
    for c in df.columns:
        if c not in result.columns: result[c] = df[c]
    return result

def action_cat(text: str) -> str:
    t = text.lower()
    if any(w in t for w in ["modifier","25 mod","51 mod","59 mod","add mod"]): return "modifier"
    if any(w in t for w in ["medical record","mr appeal","submit with mr","appeal with mr"]): return "medical_records"
    if any(w in t for w in ["adjust","write off","write-off","non-billable","non billable"]): return "adjust"
    if any(w in t for w in ["authorization","prior auth"," auth ","obtain auth"]): return "auth"
    if any(w in t for w in ["formal appeal","file appeal","appeal if"]): return "appeal"
    if any(w in t for w in ["bill patient","deductible","copay","coinsurance"]): return "bill_patient"
    if any(w in t for w in ["timely filing","filing limit"]): return "timely_filing"
    if any(w in t for w in ["resubmit","re-submit","corrected claim"]): return "resubmit"
    return "default"

def risk_lvl(days: int) -> str:
    return "HIGH" if days >= 200 else "MEDIUM" if days >= 100 else "LOW"

def conf_lbl(score: int) -> str:
    return "High" if score >= 9 else "Medium" if score >= 5 else "Low"

def ar_mult(score: int) -> float:
    for (lo, hi), m in SCORE_MULT.items():
        if lo <= score < hi: return m
    return 0.70

def notes_to_action(notes: str) -> str:
    """Feature 3: infer suggested action from prior activity notes when no denial code."""
    if not notes: return ""
    t = notes.lower()
    parts = []
    if any(w in t for w in ["modifier","mod 25","mod 51","mod 59","add mod"]):
        parts.append("Review modifier — prior notes show modifier was added previously")
    if any(w in t for w in ["medical record","mr","appeal"]):
        parts.append("Prepare Medical Records (MR) — prior notes indicate MR submission was required")
    if any(w in t for w in ["adjust","non-billable","non billable","write off","write-off"]):
        parts.append("Consult coding team — prior notes indicate code adjustment may be needed")
    if any(w in t for w in ["auth","authorization","prior auth"]):
        parts.append("Verify authorization — prior notes indicate auth was required")
    if any(w in t for w in ["bill patient","patient pay","deductible","copay"]):
        parts.append("Bill patient — prior notes indicate patient financial responsibility")
    if any(w in t for w in ["resubmit","corrected","re-filed","reprocessed"]):
        parts.append("Resubmit with corrections — prior notes indicate resubmission resolved similar claims")
    if any(w in t for w in ["timely","filing","deadline"]):
        parts.append("Check timely filing window — prior notes reference filing deadline concerns")
    return "  |  ".join(parts) if parts else \
           "Review prior activity notes — manual review recommended; no clear action pattern found"


# ══════════════════════════════════════════════════════════════════════════
#  CORE ENGINE CLASS
# ══════════════════════════════════════════════════════════════════════════

class PayerRiskEngine:
    def __init__(self, avg_claim_value: float = 850.0):
        self.training_df: Optional[pd.DataFrame] = None
        self._idx: dict = {}
        self._trained = False
        self.avg_claim_value = avg_claim_value

    # ── Train ──────────────────────────────────────────────────────────
    def train(self, source, append=False):
        df = source.copy() if isinstance(source, pd.DataFrame) \
             else load_csv_or_excel(source, TRAIN_ALIASES)
        df["days"]      = pd.to_numeric(df["days"], errors="coerce").fillna(120).astype(int)
        df["denial"]    = df["denial"].str.upper().str.strip()
        df["cpt"]       = df["cpt"].str.strip()
        df["dx"]        = df["dx"].str.upper().str.strip()
        df["payer_l"]   = df["payer"].str.lower().str.strip()
        df["payer_f"]   = df["payer_l"].apply(payer_fam)
        df["denial_p4"] = df["denial"].str[:4]
        df["cpt_p3"]    = df["cpt"].str[:3]
        df["dx_p3"]     = df["dx"].str[:3]
        df["dx_p1"]     = df["dx"].str[:1]
        self.training_df = pd.concat([self.training_df, df], ignore_index=True) \
                           if (append and self.training_df is not None) else df.reset_index(drop=True)
        self._build_idx()
        self._trained = True
        log.info(f"Trained — {len(self.training_df):,} records indexed.")

    def _build_idx(self):
        self._idx = {"d":{}, "c":{}, "p":{}}
        for i, row in self.training_df.iterrows():
            self._idx["d"].setdefault(row["denial_p4"], []).append(i)
            self._idx["c"].setdefault(row["cpt_p3"],    []).append(i)
            self._idx["p"].setdefault(row["payer_f"],   []).append(i)

    def _cands(self, denial, cpt, payer):
        s = set()
        s.update(self._idx["d"].get(denial[:4] if denial else "", []))
        s.update(self._idx["c"].get(cpt[:3]    if cpt    else "", []))
        s.update(self._idx["p"].get(payer_fam(payer),             []))
        return s or set(self.training_df.index)

    def _score_row(self, pl, pf, cpt, dxu, du, dp4, row):
        sc = 0
        if du and du == row["denial"]:    sc += 6
        elif dp4 == row["denial_p4"]:     sc += 2
        if cpt == row["cpt"]:             sc += 5
        elif cpt[:3] == row["cpt_p3"]:   sc += 2
        if pl == row["payer_l"]:          sc += 4
        elif pf == row["payer_f"]:        sc += 1
        if dxu[:3] == row["dx_p3"]:      sc += 2
        elif dxu[:1] == row["dx_p1"]:    sc += 1
        return sc

    def _best_match(self, payer, cpt, dx, denial):
        pl = payer.lower(); pf = payer_fam(payer)
        du = denial.upper(); dp4 = du[:4]
        dxu = dx.upper()
        best, bi = -1, -1
        for i in self._cands(du, cpt, payer):
            sc = self._score_row(pl, pf, cpt, dxu, du, dp4, self.training_df.loc[i])
            if sc > best: best, bi = sc, i
        return best, bi

    def _basis(self, payer, cpt, dx, denial, bi):
        r = self.training_df.loc[bi]; parts = []
        if denial.upper() == r["denial"]:           parts.append(f"Denial:{denial}")
        if cpt == r["cpt"]:                          parts.append(f"CPT:{cpt}")
        elif cpt[:3] == r["cpt_p3"]:               parts.append(f"CPT-pfx:{cpt[:3]}")
        if payer.lower() == r["payer_l"]:            parts.append(f"Payer:{payer}")
        elif payer_fam(payer) == r["payer_f"]:      parts.append(f"Family:{payer_fam(payer)}")
        if dx[:3].upper() == r["dx_p3"]:            parts.append(f"DX-pfx:{dx[:3]}")
        return "  ·  ".join(parts) if parts else "Closest available match"

    # ── Feature 1: AR Days Saved ───────────────────────────────────────
    def _ar_saved(self, action: str, score: int) -> dict:
        cat  = action_cat(action)
        ref  = AR_MAP.get(cat, AR_MAP["default"])
        mult = ar_mult(score)
        saved_adj = round(ref["saved"] * mult)
        pct = round(saved_adj / ref["baseline"] * 100, 1) if ref["baseline"] else 0
        return {
            "Action_Category":     cat.replace("_"," ").title(),
            "AR_Baseline_Days":    ref["baseline"],
            "AR_After_Action_Days":ref["after"],
            "AR_Days_Saved":       saved_adj,
            "AR_Pct_Saved":        f"{pct}%",
        }

    # ── Feature 2: Success Rate + Collection Projection ───────────────
    def _projection(self, denial: str, confidence: str, score: int, amount: float) -> dict:
        base   = SUCCESS_RATE.get(denial.upper(), SUCCESS_RATE["DEFAULT"])
        cm     = CONF_MULT.get(confidence, 0.80)
        bonus  = min(0.08, max(0, score - 5) * 0.01)
        rate   = min(0.99, round(base * cm + bonus, 4))
        collect = round(amount * rate, 2)
        writeoff = round(amount * (1 - rate), 2)
        return {
            "Est_Success_Rate":        f"{round(rate*100,1)}%",
            "Est_Collection_Amount":   collect,
            "Est_Write_Off_Risk":      writeoff,
            "_success_rate_raw":       rate,
        }

    # ── predict_single ─────────────────────────────────────────────────
    def predict_single(self, payer="", cpt="", dx="", denial="",
                       notes="", amount=None) -> dict:
        if not self._trained:
            raise RuntimeError("Call .train() first.")
        payer  = ns(payer);  cpt    = ns(cpt)
        dx     = ns(dx);     denial = ns(denial).upper()
        notes  = ns(notes)
        amt    = self.avg_claim_value
        try:
            v = float(amount)
            if not np.isnan(v) and v > 0: amt = v
        except (TypeError, ValueError): pass

        no_denial = denial in ("", "—", "N/A", "NA")

        # Score against training
        score, bi = self._best_match(payer, cpt, dx, denial)

        if bi >= 0 and score >= 0:
            row    = self.training_df.loc[bi]
            action = row["action"]
            days   = int(row["days"])
            if payer.lower() != row["payer_l"] and payer_fam(payer) != row["payer_f"]:
                days += 30
            confidence = conf_lbl(score)
            basis      = self._basis(payer, cpt, dx, denial, bi)
        else:
            dflt       = DENIAL_DEFAULTS.get(denial, {"action":"Review denial; consult coding; resubmit or appeal","days":120})
            action     = dflt["action"]
            days       = dflt["days"]
            confidence = "Low"
            basis      = "Denial code default — no training match"
            score      = 0

        # Feature 3: No-denial handling via notes
        notes_suggestion = ""
        no_denial_flag   = "No"
        if no_denial:
            no_denial_flag   = "Yes"
            notes_suggestion = notes_to_action(notes)
            if score < 5 and notes_suggestion:
                action     = notes_suggestion
                confidence = "Medium"
                basis      = "Derived from prior activity notes (no denial code provided)"

        # Feature 1: AR Days Saved
        ar = self._ar_saved(action, score)

        # Feature 2: Collection Projection
        denial_key = denial if not no_denial else "DEFAULT"
        proj = self._projection(denial_key, confidence, score, amt)

        return {
            # Input fields echoed
            "Predicted_Action":         action,
            "Notes_Derived_Suggestion": notes_suggestion,
            "No_Denial_Claim":          no_denial_flag,
            "Est_Resolution_Days":      days,
            "Risk_Level":               risk_lvl(days),
            "Confidence":               confidence,
            "Match_Score":              score,
            "Match_Basis":              basis,
            # Feature 1
            "Action_Category":          ar["Action_Category"],
            "AR_Baseline_Days":         ar["AR_Baseline_Days"],
            "AR_After_Action_Days":     ar["AR_After_Action_Days"],
            "AR_Days_Saved":            ar["AR_Days_Saved"],
            "AR_Pct_Saved":             ar["AR_Pct_Saved"],
            # Feature 2
            "Charge_Amount":            amt,
            "Est_Success_Rate":         proj["Est_Success_Rate"],
            "Est_Collection_Amount":    proj["Est_Collection_Amount"],
            "Est_Write_Off_Risk":       proj["Est_Write_Off_Risk"],
            "_success_rate_raw":        proj["_success_rate_raw"],
        }

    # ── Batch predict ──────────────────────────────────────────────────
    def predict(self, source, chunk_size=50_000) -> pd.DataFrame:
        if not self._trained: raise RuntimeError("Call .train() first.")
        if isinstance(source, pd.DataFrame):
            return self._run(source, "payer","cpt","dx","denial","notes","amount")
        path = Path(source)
        if path.suffix.lower() in [".xlsx",".xls"]:
            return self._run(load_csv_or_excel(source, INPUT_ALIASES),
                             "payer","cpt","dx","denial","notes","amount")
        # Large CSV — chunked
        parts, total, t0 = [], 0, time.time()
        for i, chunk in enumerate(pd.read_csv(source, dtype=str,
                                               chunksize=chunk_size, encoding="utf-8-sig")):
            chunk.columns = chunk.columns.str.strip()
            for std, aliases in INPUT_ALIASES.items():
                col = find_col(chunk, aliases, required=False)
                chunk[f"__{std}"] = chunk[col].fillna("").str.strip() if col \
                                     else pd.Series([""] * len(chunk))
            result = self._run(chunk,"__payer","__cpt","__dx","__denial","__notes","__amount")
            result.drop(columns=[c for c in result.columns if c.startswith("__")], inplace=True)
            parts.append(result); total += len(chunk)
            log.info(f"  Chunk {i+1}: {total:,} records ({time.time()-t0:.1f}s)")
        final = pd.concat(parts, ignore_index=True)
        log.info(f"Done — {len(final):,} in {time.time()-t0:.1f}s")
        return final

    def _run(self, df, pc, cc, dxc, dnc, ntc, amc) -> pd.DataFrame:
        rows = []
        for _, row in df.iterrows():
            rows.append(self.predict_single(
                payer=ns(row.get(pc,"")), cpt=ns(row.get(cc,"")),
                dx=ns(row.get(dxc,"")),  denial=ns(row.get(dnc,"")),
                notes=ns(row.get(ntc,"")),amount=row.get(amc,None)))
        pred = pd.DataFrame(rows)
        out  = df.copy()
        for col in pred.columns: out[col] = pred[col].values
        return out

    # ── Summary stats ──────────────────────────────────────────────────
    def summary(self, df: pd.DataFrame) -> dict:
        n   = len(df)
        rc  = df["Risk_Level"].value_counts().to_dict()
        cc2 = df["Confidence"].value_counts().to_dict()
        tch = df["Charge_Amount"].sum()
        tcl = df["Est_Collection_Amount"].sum()
        two = df["Est_Write_Off_Risk"].sum()
        nd  = int((df.get("No_Denial_Claim","No") == "Yes").sum())
        return {
            "total":         n,
            "high":          rc.get("HIGH",0),
            "medium":        rc.get("MEDIUM",0),
            "low":           rc.get("LOW",0),
            "pct_high":      round(rc.get("HIGH",0)/n*100,1),
            "avg_days":      round(df["Est_Resolution_Days"].mean(),1),
            "avg_ar_saved":  round(df["AR_Days_Saved"].mean(),1),
            "total_ar_saved":int(df["AR_Days_Saved"].sum()),
            "conf_high":     cc2.get("High",0),
            "conf_med":      cc2.get("Medium",0),
            "conf_low":      cc2.get("Low",0),
            "no_denial":     nd,
            "total_charge":  round(tch,2),
            "total_collect": round(tcl,2),
            "total_writeoff":round(two,2),
            "success_rate":  f"{round(tcl/tch*100,1)}%" if tch else "N/A",
        }

    # ── Export ─────────────────────────────────────────────────────────
    def export(self, df: pd.DataFrame, path: str, high_risk_only=False):
        data = df[df["Risk_Level"]=="HIGH"].copy() if high_risk_only else df.copy()
        # Drop internal column
        data = data.drop(columns=["_success_rate_raw"], errors="ignore")
        p = Path(path); p.parent.mkdir(parents=True, exist_ok=True)
        if p.suffix.lower() in [".xlsx",".xls"]:
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    data.to_excel(writer, sheet_name="Results", index=False)
                    ws = writer.sheets["Results"]
                    ci = {c:i+1 for i,c in enumerate(data.columns)}
                    rf = {"HIGH":("B91C1C","FFF"),"MEDIUM":("B45309","FFF"),"LOW":("166534","FFF")}
                    cf = {"High":("DCFCE7","166534"),"Medium":("FEF9C3","854D0E"),"Low":("FEE2E2","B91C1C")}
                    for r in range(2, len(data)+2):
                        rv = ws.cell(r, ci.get("Risk_Level",1)).value
                        cv = ws.cell(r, ci.get("Confidence",1)).value
                        if rv in rf and "Risk_Level" in ci:
                            c = ws.cell(r, ci["Risk_Level"])
                            c.fill=PatternFill("solid",fgColor=rf[rv][0]); c.font=Font(bold=True,color=rf[rv][1])
                        if cv in cf and "Confidence" in ci:
                            c = ws.cell(r, ci["Confidence"])
                            c.fill=PatternFill("solid",fgColor=cf[cv][0]); c.font=Font(bold=True,color=cf[cv][1])
                log.info(f"Excel saved: {path}")
            except ImportError:
                path = path.replace(".xlsx",".csv"); data.to_csv(path,index=False)
        else:
            data.to_csv(path, index=False); log.info(f"CSV saved: {len(data):,} rows → {path}")
        return path

    # ── Feature 4: HTML Dashboard ──────────────────────────────────────
    def generate_dashboard(self, df: pd.DataFrame, path: str,
                           title="Payer Claims Risk Engine — Dashboard v4"):
        s  = self.summary(df)
        p  = Path(path); p.parent.mkdir(parents=True, exist_ok=True)

        # Prep chart data
        rc  = df["Risk_Level"].value_counts().to_dict()
        cc2 = df["Confidence"].value_counts().to_dict()
        ac  = df["Action_Category"].value_counts().head(8).to_dict()

        pcol = next((c for c in ["Payer_Name","payer"] if c in df.columns), None)
        dcol = next((c for c in ["Denial_Reason","denial"] if c in df.columns), None)

        payer_data = {}
        if pcol:
            for pay, g in df.groupby(pcol):
                payer_data[str(pay)] = {
                    "H": int((g["Risk_Level"]=="HIGH").sum()),
                    "M": int((g["Risk_Level"]=="MEDIUM").sum()),
                    "L": int((g["Risk_Level"]=="LOW").sum()),
                    "ar_saved": round(g["AR_Days_Saved"].mean(),1),
                    "collect":  round(g["Est_Collection_Amount"].sum(),2),
                    "rate":     f"{round(g['_success_rate_raw'].mean()*100,1)}%",
                }

        denial_data = {}
        if dcol:
            for den, g in df.groupby(dcol):
                if str(den).strip() in ["","—","nan"]: continue
                denial_data[str(den)] = {
                    "count":   len(g),
                    "avg_days":round(g["Est_Resolution_Days"].mean(),1),
                    "ar_saved":round(g["AR_Days_Saved"].mean(),1),
                    "rate":    f"{round(g['_success_rate_raw'].mean()*100,1)}%",
                    "collect": round(g["Est_Collection_Amount"].sum(),2),
                }

        # AR saved by action category
        ar_cat = {}
        for cat, g in df.groupby("Action_Category"):
            ar_cat[str(cat)] = {
                "avg_saved": round(g["AR_Days_Saved"].mean(),1),
                "count":     len(g),
                "collect":   round(g["Est_Collection_Amount"].sum(),2),
            }

        preview_cols = [c for c in [
            "Claim_ID","Payer_Name","CPT_Code","DX_Code","Denial_Reason",
            "Predicted_Action","Est_Resolution_Days","Risk_Level","Confidence",
            "AR_Days_Saved","AR_Pct_Saved","Est_Success_Rate",
            "Est_Collection_Amount","Est_Write_Off_Risk","Action_Category",
            "No_Denial_Claim","Match_Basis"
        ] if c in df.columns]
        preview = df[preview_cols].head(200).fillna("").to_dict(orient="records")

        html = _build_dashboard_html(title, s, rc, cc2, ac,
                                     payer_data, denial_data, ar_cat, preview)
        with open(path, "w", encoding="utf-8") as f: f.write(html)
        log.info(f"Dashboard saved: {path}")
        return path


# ═══════════════════════════════════════════════════════════════════════════
#  DASHBOARD HTML GENERATOR
# ═══════════════════════════════════════════════════════════════════════════

def _build_dashboard_html(title, s, rc, cc, ac, payer_data, denial_data, ar_cat, preview):
    gen_time = time.strftime("%Y-%m-%d %H:%M")
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:system-ui,-apple-system,Arial,sans-serif;background:#f0f4f8;color:#1a1a2e;font-size:13px}}
/* HEADER */
.hdr{{background:linear-gradient(135deg,#1e3a5f 0%,#1d4ed8 100%);padding:22px 32px;display:flex;align-items:center;gap:14px}}
.hdr-logo{{width:42px;height:42px;background:rgba(255,255,255,.15);border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;font-weight:700;color:#fff;border:1px solid rgba(255,255,255,.3)}}
.hdr-t h1{{font-size:19px;font-weight:700;color:#fff;margin:0}}
.hdr-t p{{font-size:11px;color:rgba(255,255,255,.65);margin-top:2px}}
.hdr-right{{margin-left:auto;display:flex;gap:8px;flex-wrap:wrap}}
.badge{{padding:4px 12px;border-radius:20px;font-size:11px;font-weight:600;border:1px solid rgba(255,255,255,.3);color:#fff}}
/* BODY */
.body{{padding:22px 32px}}
/* STAT CARDS */
.kpi-row{{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin-bottom:18px}}
.kpi{{background:#fff;border-radius:12px;padding:16px;border:1px solid #e2e8f0;box-shadow:0 1px 4px rgba(0,0,0,.05)}}
.kpi-v{{font-size:24px;font-weight:700;margin-bottom:4px}}
.kpi-l{{font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.5px}}
.kpi-sub{{font-size:11px;margin-top:4px;font-weight:500}}
.c-blue{{color:#1d4ed8}}.c-red{{color:#b91c1c}}.c-amber{{color:#b45309}}.c-green{{color:#166534}}.c-purple{{color:#7c3aed}}
/* CARDS */
.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:18px}}
.grid3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px;margin-bottom:18px}}
.card{{background:#fff;border-radius:12px;padding:18px;border:1px solid #e2e8f0;box-shadow:0 1px 4px rgba(0,0,0,.05)}}
.card h3{{font-size:13px;font-weight:600;color:#374151;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #f1f5f9}}
.ch{{position:relative;height:210px}}
/* PILLS */
.pill{{display:inline-block;padding:2px 8px;border-radius:20px;font-size:10px;font-weight:700}}
.p-h{{background:#fee2e2;color:#b91c1c}}.p-m{{background:#fef9c3;color:#854d0e}}
.p-l{{background:#dcfce7;color:#166534}}.p-hi{{background:#dcfce7;color:#166534}}
.p-md{{background:#fef9c3;color:#854d0e}}.p-lo{{background:#fee2e2;color:#b91c1c}}
.p-b{{background:#dbeafe;color:#1e40af}}.p-g{{background:#f0fdf4;color:#166534}}
.p-ind{{background:#ede9fe;color:#5b21b6}}
/* TABS */
.tabs{{display:flex;border-bottom:1px solid #e2e8f0;margin-bottom:14px}}
.tab{{padding:8px 16px;font-size:12px;font-weight:500;background:none;border:none;border-bottom:2px solid transparent;color:#64748b;cursor:pointer;margin-bottom:-1px}}
.tab.on{{color:#1d4ed8;border-bottom-color:#1d4ed8;font-weight:600}}
.panel{{display:none}}.panel.on{{display:block}}
/* TABLE */
.tbl-wrap{{max-height:360px;overflow-y:auto;border:1px solid #e2e8f0;border-radius:8px}}
table{{width:100%;border-collapse:collapse;font-size:11px}}
th{{padding:8px;text-align:left;color:#6b7280;border-bottom:1px solid #e2e8f0;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.3px;background:#f8fafc;position:sticky;top:0;z-index:1}}
td{{padding:7px 8px;border-bottom:1px solid #f1f5f9;vertical-align:middle}}
tr:hover td{{background:#fafbfc}}
/* PAYER BARS */
.pbar-row{{display:flex;align-items:center;gap:8px;margin-bottom:8px}}
.pbar-name{{width:110px;font-size:11px;font-weight:500;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:#374151}}
.pbar-track{{flex:1;height:14px;background:#f1f5f9;border-radius:7px;overflow:hidden;display:flex}}
.ph{{height:14px;background:#b91c1c}}.pm{{height:14px;background:#b45309}}.pl{{height:14px;background:#166534}}
.pbar-n{{font-size:10px;color:#6b7280;width:24px;text-align:right}}
/* DENIAL TABLE */
.dt-row{{display:grid;grid-template-columns:80px 1fr 65px 65px 65px 90px;gap:8px;align-items:center;padding:6px 0;border-bottom:1px solid #f1f5f9;font-size:11px}}
.dt-hdr{{font-size:9px;color:#6b7280;font-weight:600;text-transform:uppercase;border-bottom:1px solid #e2e8f0;padding-bottom:7px}}
.mini{{height:7px;background:#e5e7eb;border-radius:4px;overflow:hidden}}
.mini-f{{height:7px;background:#3b82f6;border-radius:4px}}
/* AR BARS */
.ar-row{{display:flex;align-items:center;gap:12px;padding:7px 10px;background:#f8fafc;border-radius:8px;margin-bottom:7px;border:1px solid #e2e8f0}}
.ar-cat{{font-size:12px;font-weight:500;color:#374151;width:150px;flex-shrink:0}}
.ar-track{{flex:1;height:10px;background:#e5e7eb;border-radius:5px;overflow:hidden}}
.ar-fill{{height:10px;background:#1d4ed8;border-radius:5px}}
.ar-val{{font-size:13px;font-weight:700;color:#1d4ed8;width:80px;text-align:right;flex-shrink:0}}
/* FILTERS */
.filters{{display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap;align-items:center}}
.filters select,.filters input{{padding:5px 10px;border:1px solid #d1d5db;border-radius:6px;font-size:12px;font-family:inherit;background:#fff}}
.btn{{padding:5px 14px;border-radius:6px;font-size:12px;font-weight:500;cursor:pointer;border:1px solid #d1d5db;background:#f9fafb;font-family:inherit}}
.btn-dl{{background:#166534;color:#fff;border-color:#166534}}
.btn-dl:hover{{background:#15803d}}
.row-count{{font-size:11px;color:#6b7280;margin-left:auto}}
/* COLLECTION SUB-KPIS */
.ckpi-row{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:14px}}
.ckpi{{text-align:center;padding:12px;background:#f8fafc;border-radius:8px;border:1px solid #e2e8f0}}
.ckpi-v{{font-size:19px;font-weight:700;margin-bottom:2px}}
.ckpi-l{{font-size:10px;color:#6b7280;text-transform:uppercase;letter-spacing:.4px}}
.footer{{text-align:center;padding:14px 32px;color:#94a3b8;font-size:11px;border-top:1px solid #e2e8f0;background:#fff;margin-top:20px}}
</style>
</head>
<body>
<div class="hdr">
  <div class="hdr-logo">RE</div>
  <div class="hdr-t">
    <h1>{title}</h1>
    <p>AR Days Saved · Success Rate · Collection Projection · No-Denial Analysis · Auto-Generated Dashboard</p>
  </div>
  <div class="hdr-right">
    <span class="badge">● {s['total']:,} Claims Analysed</span>
    <span class="badge" style="background:rgba(22,101,52,.35)">{s['success_rate']} Recovery Rate</span>
    <span class="badge" style="background:rgba(124,58,237,.35)">{s['avg_ar_saved']:.0f} Avg AR Days Saved</span>
  </div>
</div>

<div class="body">

<!-- KPI ROW -->
<div class="kpi-row">
  <div class="kpi"><div class="kpi-v c-blue">{s['total']:,}</div><div class="kpi-l">Total Claims</div><div class="kpi-sub" style="color:#64748b">{s['no_denial']} no-denial claims</div></div>
  <div class="kpi"><div class="kpi-v c-red">{s['high']:,}</div><div class="kpi-l">HIGH Risk</div><div class="kpi-sub c-red">{s['pct_high']}% of total</div></div>
  <div class="kpi"><div class="kpi-v c-amber">{s['avg_ar_saved']:.0f} days</div><div class="kpi-l">Avg AR Days Saved</div><div class="kpi-sub c-green">↓ vs manual baseline</div></div>
  <div class="kpi"><div class="kpi-v c-green">${s['total_collect']:,.0f}</div><div class="kpi-l">Est. Total Collection</div><div class="kpi-sub" style="color:#64748b">{s['success_rate']} success rate</div></div>
  <div class="kpi"><div class="kpi-v c-purple">${s['total_writeoff']:,.0f}</div><div class="kpi-l">Est. Write-Off Risk</div><div class="kpi-sub c-red">Requires escalation</div></div>
</div>

<!-- CHARTS ROW 1 -->
<div class="grid3">
  <div class="card"><h3>📊 Risk Level Distribution</h3><div class="ch"><canvas id="riskC"></canvas></div></div>
  <div class="card"><h3>🎯 Prediction Confidence</h3><div class="ch"><canvas id="confC"></canvas></div></div>
  <div class="card"><h3>⚡ Action Categories</h3><div class="ch"><canvas id="actionC"></canvas></div></div>
</div>

<!-- PAYER + DENIAL -->
<div class="grid2">
  <div class="card">
    <h3>🏥 Payer Risk Breakdown (H / M / L)</h3>
    <div id="payerBars" style="max-height:260px;overflow-y:auto;padding-right:4px"></div>
  </div>
  <div class="card">
    <h3>🔍 Denial Code Analysis</h3>
    <div class="dt-row dt-hdr"><div>Code</div><div>Volume</div><div>Avg Days</div><div>AR Saved</div><div>Success</div><div>Collection</div></div>
    <div id="denialRows" style="max-height:220px;overflow-y:auto"></div>
  </div>
</div>

<!-- AR DAYS SAVED BY CATEGORY -->
<div class="card" style="margin-bottom:18px">
  <h3>⏱ AR Days Saved by Action Category</h3>
  <div id="arBars"></div>
</div>

<!-- TABBED DETAIL -->
<div class="card">
  <div class="tabs">
    <button class="tab on" onclick="goTab('claims',this)">📋 Claims Detail</button>
    <button class="tab" onclick="goTab('collection',this)">💰 Collection Summary</button>
    <button class="tab" onclick="goTab('nodenial',this)">🔎 No-Denial Claims</button>
  </div>

  <!-- CLAIMS TAB -->
  <div class="panel on" id="tab-claims">
    <div class="filters">
      <select id="fRisk" onchange="filter()"><option value="">All Risk</option><option>HIGH</option><option>MEDIUM</option><option>LOW</option></select>
      <select id="fConf" onchange="filter()"><option value="">All Confidence</option><option>High</option><option>Medium</option><option>Low</option></select>
      <select id="fND" onchange="filter()"><option value="">All Claims</option><option value="Yes">No-Denial Only</option><option value="No">Denial Claims</option></select>
      <input id="fSearch" placeholder="Search payer / CPT / denial / action..." oninput="filter()" style="min-width:220px">
      <button class="btn btn-dl" onclick="dlCSV()">⬇ Download CSV</button>
      <span class="row-count" id="rowCount"></span>
    </div>
    <div class="tbl-wrap">
      <table>
        <thead><tr>
          <th>Claim ID</th><th>Payer</th><th>CPT</th><th>DX</th><th>Denial</th>
          <th>Predicted Action</th><th>Est Days</th><th>Risk</th><th>Conf</th>
          <th>AR Saved</th><th>Success %</th><th>Collection $</th><th>Category</th><th>No-Denial</th>
        </tr></thead>
        <tbody id="tBody"></tbody>
      </table>
    </div>
  </div>

  <!-- COLLECTION TAB -->
  <div class="panel" id="tab-collection">
    <div class="ckpi-row">
      <div class="ckpi"><div class="ckpi-v c-blue">${s['total_charge']:,.2f}</div><div class="ckpi-l">Total Charged</div></div>
      <div class="ckpi"><div class="ckpi-v c-green">${s['total_collect']:,.2f}</div><div class="ckpi-l">Est. Collectible</div></div>
      <div class="ckpi"><div class="ckpi-v c-red">${s['total_writeoff']:,.2f}</div><div class="ckpi-l">Est. Write-Off Risk</div></div>
    </div>
    <div class="ch" style="max-width:600px;margin:0 auto"><canvas id="collC"></canvas></div>
  </div>

  <!-- NO-DENIAL TAB -->
  <div class="panel" id="tab-nodenial">
    <p style="font-size:12px;color:#64748b;margin-bottom:12px">Claims without a denial code — action derived from prior activity notes.</p>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th>Claim ID</th><th>Payer</th><th>CPT</th><th>DX</th><th>Predicted Action</th><th>Est Days</th><th>Risk</th><th>AR Saved</th><th>Success %</th></tr></thead>
        <tbody id="ndBody"></tbody>
      </table>
    </div>
  </div>
</div>

</div>
<div class="footer">Payer Claims Risk Engine v4 · Generated {gen_time} · {s['total']:,} claims analysed · All processing local — no data shared</div>

<script>
const RC={json.dumps(rc)};
const CC={json.dumps(cc)};
const AC={json.dumps(ac)};
const PD={json.dumps(payer_data)};
const DD={json.dumps(denial_data)};
const AR={json.dumps(ar_cat)};
const ROWS={json.dumps(preview)};
let filtered=[...ROWS];

// ── Charts ──────────────────────────────────────────────────────────
new Chart(document.getElementById('riskC'),{{type:'doughnut',
  data:{{labels:Object.keys(RC),datasets:[{{data:Object.values(RC),
    backgroundColor:['#b91c1c','#b45309','#166534'],borderColor:'#fff',borderWidth:3}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'right',labels:{{font:{{size:11}}}}}}}}}}
}});

new Chart(document.getElementById('confC'),{{type:'bar',
  data:{{labels:Object.keys(CC),datasets:[{{data:Object.values(CC),
    backgroundColor:['#166534','#b45309','#b91c1c'],borderRadius:6,borderSkipped:false}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}}}},
    scales:{{y:{{beginAtZero:true,grid:{{color:'#f1f5f9'}}}},x:{{grid:{{display:false}}}}}}}}
}});

new Chart(document.getElementById('actionC'),{{type:'bar',
  data:{{labels:Object.keys(AC),datasets:[{{data:Object.values(AC),
    backgroundColor:'#3b82f6',borderRadius:4,borderSkipped:false}}]}},
  options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}}}},
    scales:{{x:{{beginAtZero:true,grid:{{color:'#f1f5f9'}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:10}}}}}}}}}}
}});

new Chart(document.getElementById('collC'),{{type:'bar',
  data:{{labels:['Total Charged','Est. Collection','Est. Write-Off'],
    datasets:[{{data:[{s['total_charge']},{s['total_collect']},{s['total_writeoff']}],
      backgroundColor:['#1d4ed8','#166534','#b91c1c'],borderRadius:8,borderSkipped:false}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>'$'+ctx.raw.toLocaleString('en-US',{{minimumFractionDigits:2}})}}}}}},
    scales:{{y:{{beginAtZero:true,ticks:{{callback:v=>'$'+v.toLocaleString()}},grid:{{color:'#f1f5f9'}}}},x:{{grid:{{display:false}}}}}}}}
}});

// ── Payer bars ──────────────────────────────────────────────────────
const pbWrap=document.getElementById('payerBars');
Object.entries(PD).sort((a,b)=>(b[1].H+b[1].M+b[1].L)-(a[1].H+a[1].M+a[1].L)).forEach(([name,d])=>{{
  const tot=(d.H+d.M+d.L)||1;
  pbWrap.innerHTML+=`<div class="pbar-row">
    <div class="pbar-name" title="${{name}}">${{name}}</div>
    <div class="pbar-track">
      <div class="ph" style="width:${{d.H/tot*100}}%"></div>
      <div class="pm" style="width:${{d.M/tot*100}}%"></div>
      <div class="pl" style="width:${{d.L/tot*100}}%"></div>
    </div>
    <div class="pbar-n">${{tot}}</div>
  </div>`;
}});

// ── Denial rows ─────────────────────────────────────────────────────
const maxCnt=Math.max(...Object.values(DD).map(d=>d.count),1);
const drWrap=document.getElementById('denialRows');
Object.entries(DD).sort((a,b)=>b[1].count-a[1].count).forEach(([code,d])=>{{
  drWrap.innerHTML+=`<div class="dt-row">
    <div><span class="pill p-h">${{code}}</span></div>
    <div><div class="mini"><div class="mini-f" style="width:${{d.count/maxCnt*100}}%"></div></div>
         <span style="font-size:10px;color:#6b7280">${{d.count}}</span></div>
    <div style="font-weight:600">${{d.avg_days}}d</div>
    <div style="font-weight:600;color:#1d4ed8">${{d.ar_saved}}d</div>
    <div style="font-weight:600;color:#166534">${{d.rate}}</div>
    <div style="font-weight:600">$${{d.collect.toLocaleString('en-US',{{minimumFractionDigits:2}})}}</div>
  </div>`;
}});

// ── AR bars ─────────────────────────────────────────────────────────
const maxAR=Math.max(...Object.values(AR).map(v=>v.avg_saved),1);
const arWrap=document.getElementById('arBars');
Object.entries(AR).sort((a,b)=>b[1].avg_saved-a[1].avg_saved).forEach(([cat,d])=>{{
  arWrap.innerHTML+=`<div class="ar-row">
    <div class="ar-cat">${{cat}} <span style="font-size:10px;color:#6b7280">(${{d.count}} claims)</span></div>
    <div class="ar-track"><div class="ar-fill" style="width:${{d.avg_saved/maxAR*100}}%"></div></div>
    <div class="ar-val">${{d.avg_saved}} days</div>
  </div>`;
}});

// ── Table rendering ─────────────────────────────────────────────────
function riskClass(r){{return r==='HIGH'?'p-h':r==='MEDIUM'?'p-m':'p-l'}}
function confClass(c){{return c==='High'?'p-hi':c==='Medium'?'p-md':'p-lo'}}

function renderTable(rows, tbodyId){{
  const tb=document.getElementById(tbodyId);
  tb.innerHTML=rows.map(r=>`<tr>
    <td style="font-family:monospace;font-size:10px">${{r.Claim_ID||'—'}}</td>
    <td><strong>${{r.Payer_Name||r.payer||'—'}}</strong></td>
    <td><span class="pill p-b">${{r.CPT_Code||r.cpt||'—'}}</span></td>
    <td style="font-family:monospace;font-size:10px">${{r.DX_Code||r.dx||'—'}}</td>
    <td><span class="pill p-h">${{r.Denial_Reason||r.denial||'—'}}</span></td>
    <td style="max-width:200px;white-space:normal;font-size:10px">${{r.Predicted_Action||'—'}}</td>
    <td style="text-align:center;font-weight:600">${{r.Est_Resolution_Days||'—'}}</td>
    <td><span class="pill ${{riskClass(r.Risk_Level)}}">${{r.Risk_Level||'—'}}</span></td>
    <td><span class="pill ${{confClass(r.Confidence)}}">${{r.Confidence||'—'}}</span></td>
    <td style="text-align:center;font-weight:700;color:#1d4ed8">${{r.AR_Days_Saved||'—'}}</td>
    <td style="text-align:center;font-weight:600;color:#166534">${{r.Est_Success_Rate||'—'}}</td>
    <td style="text-align:right;font-weight:600">$${{(r.Est_Collection_Amount||0).toLocaleString('en-US',{{minimumFractionDigits:2}})}}</td>
    <td><span class="pill p-ind">${{r.Action_Category||'—'}}</span></td>
    <td style="text-align:center">${{r.No_Denial_Claim==='Yes'?'<span class="pill p-g">Yes</span>':'No'}}</td>
  </tr>`).join('');
}}

function renderNDTable(rows){{
  const tb=document.getElementById('ndBody');
  tb.innerHTML=rows.filter(r=>r.No_Denial_Claim==='Yes').map(r=>`<tr>
    <td style="font-family:monospace;font-size:10px">${{r.Claim_ID||'—'}}</td>
    <td><strong>${{r.Payer_Name||'—'}}</strong></td>
    <td><span class="pill p-b">${{r.CPT_Code||'—'}}</span></td>
    <td style="font-family:monospace;font-size:10px">${{r.DX_Code||'—'}}</td>
    <td style="max-width:220px;white-space:normal;font-size:10px">${{r.Predicted_Action||'—'}}</td>
    <td style="text-align:center;font-weight:600">${{r.Est_Resolution_Days||'—'}}</td>
    <td><span class="pill ${{riskClass(r.Risk_Level)}}">${{r.Risk_Level||'—'}}</span></td>
    <td style="text-align:center;font-weight:700;color:#1d4ed8">${{r.AR_Days_Saved||'—'}}</td>
    <td style="text-align:center;font-weight:600;color:#166534">${{r.Est_Success_Rate||'—'}}</td>
  </tr>`).join('');
}}

function filter(){{
  const risk=document.getElementById('fRisk').value;
  const conf=document.getElementById('fConf').value;
  const nd=document.getElementById('fND').value;
  const q=document.getElementById('fSearch').value.toLowerCase();
  filtered=ROWS.filter(r=>{{
    return (!risk||r.Risk_Level===risk)&&
           (!conf||r.Confidence===conf)&&
           (!nd||r.No_Denial_Claim===nd)&&
           (!q||(r.Payer_Name||'').toLowerCase().includes(q)||
                (r.CPT_Code||'').toLowerCase().includes(q)||
                (r.Denial_Reason||'').toLowerCase().includes(q)||
                (r.Predicted_Action||'').toLowerCase().includes(q));
  }});
  renderTable(filtered,'tBody');
  document.getElementById('rowCount').textContent=filtered.length+' rows shown';
}}

function dlCSV(){{
  if(!filtered.length){{alert('No data.');return;}}
  const keys=Object.keys(filtered[0]).filter(k=>!k.startsWith('_'));
  const csv=[keys.join(','),...filtered.map(r=>keys.map(k=>{{
    const v=r[k]===null||r[k]===undefined?'':String(r[k]);
    return v.includes(',')||v.includes('"')?'"'+v.replace(/"/g,'""')+'"':v;
  }}).join(','))].join('\\n');
  const a=document.createElement('a');
  a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);
  a.download='payer_risk_results.csv';a.click();
}}

function goTab(id,btn){{
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('on'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('on'));
  document.getElementById('tab-'+id).classList.add('on');
  btn.classList.add('on');
}}

renderTable(ROWS,'tBody');
renderNDTable(ROWS);
document.getElementById('rowCount').textContent=ROWS.length+' rows shown';
</script>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="Payer Claims Risk Engine v4")
    ap.add_argument("--train",          required=True,  help="Training data CSV/Excel")
    ap.add_argument("--input",          required=True,  help="Input claims CSV/Excel")
    ap.add_argument("--output",         required=True,  help="Output path .csv or .xlsx")
    ap.add_argument("--dashboard",      default="",     help="Output HTML dashboard path")
    ap.add_argument("--avg-claim",      type=float, default=850.0, help="Default claim value")
    ap.add_argument("--chunk",          type=int, default=50_000)
    ap.add_argument("--high-risk-only", action="store_true")
    ap.add_argument("--quiet",          action="store_true")
    args = ap.parse_args()
    if args.quiet: logging.getLogger().setLevel(logging.WARNING)

    engine = PayerRiskEngine(avg_claim_value=args.avg_claim)
    engine.train(args.train)

    t0 = time.time()
    results = engine.predict(args.input, chunk_size=args.chunk)
    elapsed = time.time() - t0

    out = engine.export(results, args.output, high_risk_only=args.high_risk_only)
    print(f"\n  Main output  : {out}")

    if not args.high_risk_only:
        p  = Path(args.output)
        hr = p.parent / f"{p.stem}_HIGH_RISK{p.suffix}"
        engine.export(results, str(hr), high_risk_only=True)
        print(f"  High-risk    : {hr}")

    dash = args.dashboard or str(Path(args.output).parent / "dashboard.html")
    engine.generate_dashboard(results, dash)
    print(f"  Dashboard    : {dash}")

    s = engine.summary(results)
    print(f"\n{'='*58}")
    print("  PAYER RISK ENGINE v4  —  Results Summary")
    print(f"{'='*58}")
    print(f"  Total claims        : {s['total']:>10,}")
    print(f"  Processing time     : {elapsed:>10.2f}s")
    print(f"  No-denial claims    : {s['no_denial']:>10,}")
    print(f"  HIGH risk           : {s['high']:>10,}  ({s['pct_high']}%)")
    print(f"  MEDIUM risk         : {s['medium']:>10,}")
    print(f"  LOW risk            : {s['low']:>10,}")
    print(f"  Avg AR days saved   : {s['avg_ar_saved']:>10.1f}")
    print(f"  Total AR days saved : {s['total_ar_saved']:>10,}")
    print(f"  Est. collection     : ${s['total_collect']:>10,.2f}")
    print(f"  Est. write-off risk : ${s['total_writeoff']:>10,.2f}")
    print(f"  Overall success rate: {s['success_rate']:>10}")
    print(f"{'='*58}\n")


if __name__ == "__main__":
    main()
